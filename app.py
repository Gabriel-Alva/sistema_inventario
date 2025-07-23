from fastapi import FastAPI, Request, Form, Query
from fastapi.responses import RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
import uvicorn
import os

# Importar funciones de la base de datos
from db.models import (
    insertar_producto, insertar_entrada, insertar_venta, get_connection, consultar_productos_bajos,
    total_usuarios, total_categorias, total_productos, total_ventas, productos_mas_vendidos, ultimas_ventas, productos_recientes,
    listar_categorias, agregar_categoria, editar_categoria, eliminar_categoria,
    listar_proveedores, agregar_proveedor, editar_proveedor, eliminar_proveedor,
    total_proveedores, registrar_usuario
)

# Importar librerías para exportar a Excel y PDF
import io
import openpyxl
from fpdf import FPDF
from psycopg2 import errors

# Crear la app
app = FastAPI()

# Montar archivos estáticos (CSS, JS, imágenes)
static_dir = os.path.join(os.path.dirname(__file__), 'static')
if os.path.exists(static_dir):
    app.mount("/static", StaticFiles(directory=static_dir), name="static")

# Configurar Jinja2Templates
templates = Jinja2Templates(directory=os.path.join(os.path.dirname(__file__), 'templates'))

# Rutas principales
@app.get("/")
def root():
    return RedirectResponse(url="/login")

@app.get("/login")
def login_get(request: Request):
    return templates.TemplateResponse("login.html", {"request": request})

@app.post("/login")
def login_post(request: Request, username: str = Form(...), password: str = Form(...)):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT * FROM usuarios WHERE username=%s AND password=%s", (username, password))
    user = cur.fetchone()
    cur.close()
    conn.close()
    if user:
        response = RedirectResponse(url="/dashboard", status_code=303)
        response.set_cookie(key="usuario", value=username)
        return response
    else:
        return RedirectResponse(url="/login?error=1", status_code=303)

@app.get("/dashboard")
def dashboard(request: Request):
    resumen = {
        'usuarios': total_usuarios(),
        'categorias': total_categorias(),
        'proveedores': total_proveedores(),
        'productos': total_productos(),
        'ventas': total_ventas(),
        'mas_vendidos': productos_mas_vendidos(),
        'ultimas_ventas': ultimas_ventas(),
        'productos_recientes': productos_recientes()
    }
    return templates.TemplateResponse("dashboard.html", {"request": request, "resumen": resumen})

@app.get("/productos")
def productos(request: Request, msg: str = Query(None), filtro: str = Query("")):
    conn = get_connection()
    cur = conn.cursor()
    if filtro:
        cur.execute("SELECT id_producto, nombre, descripcion, precio_venta, stock FROM productos WHERE LOWER(nombre) LIKE %s ORDER BY id_producto;", (f"%{filtro.lower()}%",))
    else:
        cur.execute("SELECT id_producto, nombre, descripcion, precio_venta, stock FROM productos ORDER BY id_producto;")
    productos = [
        {
            'id_producto': row[0],
            'nombre': row[1],
            'descripcion': row[2],
            'precio_venta': row[3],
            'stock': row[4]
        } for row in cur.fetchall()
    ]
    categorias = listar_categorias()
    proveedores = listar_proveedores()
    cur.close()
    conn.close()
    return templates.TemplateResponse("productos.html", {"request": request, "productos": productos, "msg": msg, "filtro": filtro, "categorias": categorias, "proveedores": proveedores})

@app.post("/productos/nuevo")
def productos_nuevo(
    nombre: str = Form(...),
    descripcion: str = Form(""),
    precio_venta: float = Form(...),
    precio_compra: float = Form(0),
    stock: int = Form(...),
    stock_minimo: int = Form(...)
):
    insertar_producto(nombre, descripcion, precio_venta, precio_compra, stock, stock_minimo)
    return RedirectResponse(url="/productos?msg=Producto+agregado+con+éxito", status_code=303)

@app.post("/productos/editar/{id_producto}")
def editar_producto(id_producto: int, nombre: str = Form(...), descripcion: str = Form(""), precio_venta: float = Form(...), precio_compra: float = Form(0), stock: int = Form(...), stock_minimo: int = Form(...)):
    # Si precio_compra es 0 o vacío, enviar None
    precio_compra_db = precio_compra if precio_compra and precio_compra > 0 else None
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("""
        UPDATE productos SET nombre=%s, descripcion=%s, precio_venta=%s, precio_compra=%s, stock=%s, stock_minimo=%s WHERE id_producto=%s
    """, (nombre, descripcion, precio_venta, precio_compra_db, stock, stock_minimo, id_producto))
    conn.commit()
    cur.close()
    conn.close()
    return RedirectResponse(url="/productos?msg=Producto+editado+con+éxito", status_code=303)

@app.post("/productos/eliminar/{id_producto}")
def eliminar_producto(id_producto: int):
    conn = get_connection()
    cur = conn.cursor()
    try:
        cur.execute("DELETE FROM productos WHERE id_producto=%s", (id_producto,))
        conn.commit()
        msg = "Producto eliminado con éxito"
    except Exception as e:
        conn.rollback()
        if hasattr(e, 'pgcode') and e.pgcode == '23503':
            msg = "No se puede eliminar el producto porque tiene ventas asociadas."
        else:
            msg = "Error al eliminar el producto."
    finally:
        cur.close()
        conn.close()
    return RedirectResponse(url=f"/productos?msg={msg}", status_code=303)

@app.get("/productos/exportar/excel")
def exportar_productos_excel(filtro: str = ""):
    conn = get_connection()
    cur = conn.cursor()
    if filtro:
        cur.execute("SELECT id_producto, nombre, descripcion, precio_venta, stock FROM productos WHERE LOWER(nombre) LIKE %s ORDER BY id_producto;", (f"%{filtro.lower()}%",))
    else:
        cur.execute("SELECT id_producto, nombre, descripcion, precio_venta, stock FROM productos ORDER BY id_producto;")
    productos = cur.fetchall()
    cur.close()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ID", "Nombre", "Descripción", "Precio Venta", "Stock"])
    for row in productos:
        ws.append(row)
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=productos.xlsx"})

@app.get("/productos/exportar/pdf")
def exportar_productos_pdf(filtro: str = ""):
    conn = get_connection()
    cur = conn.cursor()
    if filtro:
        cur.execute("SELECT id_producto, nombre, descripcion, precio_venta, stock FROM productos WHERE LOWER(nombre) LIKE %s ORDER BY id_producto;", (f"%{filtro.lower()}%",))
    else:
        cur.execute("SELECT id_producto, nombre, descripcion, precio_venta, stock FROM productos ORDER BY id_producto;")
    productos = cur.fetchall()
    cur.close()
    conn.close()
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(0, 10, "Listado de Productos", ln=True, align="C")
    pdf.ln(5)
    pdf.set_font("Arial", size=10)
    pdf.cell(15, 8, "ID", 1)
    pdf.cell(40, 8, "Nombre", 1)
    pdf.cell(50, 8, "Descripción", 1)
    pdf.cell(30, 8, "Precio Venta", 1)
    pdf.cell(20, 8, "Stock", 1)
    pdf.ln()
    for row in productos:
        pdf.cell(15, 8, str(row[0]), 1)
        pdf.cell(40, 8, str(row[1]), 1)
        pdf.cell(50, 8, str(row[2]), 1)
        pdf.cell(30, 8, str(row[3]), 1)
        pdf.cell(20, 8, str(row[4]), 1)
        pdf.ln()
    stream = io.BytesIO(pdf.output(dest='S').encode('latin1'))
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=productos.pdf"})

@app.get("/entradas")
def entradas(request: Request, msg: str = Query(None), filtro: str = Query("")):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id_producto, nombre FROM productos ORDER BY nombre;")
    productos = [
        {'id_producto': row[0], 'nombre': row[1]} for row in cur.fetchall()
    ]
    if filtro:
        cur.execute("""
            SELECT e.id_entrada, e.id_producto, p.nombre, e.cantidad, e.fecha_entrada
            FROM entradas e
            JOIN productos p ON e.id_producto = p.id_producto
            WHERE LOWER(p.nombre) LIKE %s
            ORDER BY e.id_entrada DESC
        """, (f"%{filtro.lower()}%",))
    else:
        cur.execute("""
            SELECT e.id_entrada, e.id_producto, p.nombre, e.cantidad, e.fecha_entrada
            FROM entradas e
            JOIN productos p ON e.id_producto = p.id_producto
            ORDER BY e.id_entrada DESC
        """)
    entradas = [
        {
            'id_entrada': row[0],
            'id_producto': row[1],
            'nombre_producto': row[2],
            'cantidad': row[3],
            'fecha_entrada': row[4].strftime('%Y-%m-%d %H:%M') if row[4] else ''
        } for row in cur.fetchall()
    ]
    cur.close()
    conn.close()
    return templates.TemplateResponse("entradas.html", {"request": request, "productos": productos, "entradas": entradas, "msg": msg, "filtro": filtro})

@app.post("/entradas/editar/{id_entrada}")
def editar_entrada(id_entrada: int, id_producto: int = Form(...), cantidad: int = Form(...)):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE entradas SET id_producto=%s, cantidad=%s WHERE id_entrada=%s", (id_producto, cantidad, id_entrada))
    conn.commit()
    cur.close()
    conn.close()
    return RedirectResponse(url="/entradas?msg=Entrada+editada+con+éxito", status_code=303)

@app.post("/entradas/eliminar/{id_entrada}")
def eliminar_entrada(id_entrada: int):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM entradas WHERE id_entrada=%s", (id_entrada,))
    conn.commit()
    cur.close()
    conn.close()
    return RedirectResponse(url="/entradas?msg=Entrada+eliminada+con+éxito", status_code=303)

@app.post("/entradas/nueva")
def registrar_entrada(id_producto: int = Form(...), cantidad: int = Form(...)):
    insertar_entrada(id_producto, cantidad)
    return RedirectResponse(url="/entradas?msg=Entrada+registrada+con+éxito", status_code=303)

@app.get("/entradas/exportar/excel")
def exportar_entradas_excel(filtro: str = ""):
    conn = get_connection()
    cur = conn.cursor()
    if filtro:
        cur.execute("""
            SELECT e.id_entrada, p.nombre, e.cantidad, e.fecha_entrada
            FROM entradas e
            JOIN productos p ON e.id_producto = p.id_producto
            WHERE LOWER(p.nombre) LIKE %s
            ORDER BY e.id_entrada DESC
        """, (f"%{filtro.lower()}%",))
    else:
        cur.execute("""
            SELECT e.id_entrada, p.nombre, e.cantidad, e.fecha_entrada
            FROM entradas e
            JOIN productos p ON e.id_producto = p.id_producto
            ORDER BY e.id_entrada DESC
        """)
    entradas = cur.fetchall()
    cur.close()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ID", "Producto", "Cantidad", "Fecha"])
    for row in entradas:
        ws.append([row[0], row[1], row[2], row[3].strftime('%Y-%m-%d %H:%M') if row[3] else ''])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=entradas.xlsx"})

@app.get("/entradas/exportar/pdf")
def exportar_entradas_pdf(filtro: str = ""):
    conn = get_connection()
    cur = conn.cursor()
    if filtro:
        cur.execute("""
            SELECT e.id_entrada, p.nombre, e.cantidad, e.fecha_entrada
            FROM entradas e
            JOIN productos p ON e.id_producto = p.id_producto
            WHERE LOWER(p.nombre) LIKE %s
            ORDER BY e.id_entrada DESC
        """, (f"%{filtro.lower()}%",))
    else:
        cur.execute("""
            SELECT e.id_entrada, p.nombre, e.cantidad, e.fecha_entrada
            FROM entradas e
            JOIN productos p ON e.id_producto = p.id_producto
            ORDER BY e.id_entrada DESC
        """)
    entradas = cur.fetchall()
    cur.close()
    conn.close()
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(0, 10, "Listado de Entradas", ln=True, align="C")
    pdf.ln(5)
    pdf.set_font("Arial", size=10)
    pdf.cell(15, 8, "ID", 1)
    pdf.cell(50, 8, "Producto", 1)
    pdf.cell(25, 8, "Cantidad", 1)
    pdf.cell(40, 8, "Fecha", 1)
    pdf.ln()
    for row in entradas:
        pdf.cell(15, 8, str(row[0]), 1)
        pdf.cell(50, 8, str(row[1]), 1)
        pdf.cell(25, 8, str(row[2]), 1)
        pdf.cell(40, 8, row[3].strftime('%Y-%m-%d %H:%M') if row[3] else '', 1)
        pdf.ln()
    stream = io.BytesIO(pdf.output(dest='S').encode('latin1'))
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=entradas.pdf"})

@app.get("/ventas")
def ventas(request: Request, msg: str = Query(None), filtro: str = Query("")):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("SELECT id_producto, nombre FROM productos ORDER BY nombre;")
    productos = [
        {'id_producto': row[0], 'nombre': row[1]} for row in cur.fetchall()
    ]
    if filtro:
        cur.execute("""
            SELECT v.id_venta, v.id_producto, p.nombre, v.cantidad, v.precio_total, v.fecha_venta
            FROM ventas v
            JOIN productos p ON v.id_producto = p.id_producto
            WHERE LOWER(p.nombre) LIKE %s
            ORDER BY v.id_venta DESC
        """, (f"%{filtro.lower()}%",))
    else:
        cur.execute("""
            SELECT v.id_venta, v.id_producto, p.nombre, v.cantidad, v.precio_total, v.fecha_venta
            FROM ventas v
            JOIN productos p ON v.id_producto = p.id_producto
            ORDER BY v.id_venta DESC
        """)
    ventas = [
        {
            'id_venta': row[0],
            'id_producto': row[1],
            'nombre_producto': row[2],
            'cantidad': row[3],
            'precio_total': row[4],
            'fecha_venta': row[5].strftime('%Y-%m-%d %H:%M') if row[5] else ''
        } for row in cur.fetchall()
    ]
    cur.close()
    conn.close()
    return templates.TemplateResponse("ventas.html", {"request": request, "productos": productos, "ventas": ventas, "msg": msg, "filtro": filtro})

@app.post("/ventas/editar/{id_venta}")
def editar_venta(id_venta: int, id_producto: int = Form(...), cantidad: int = Form(...), precio_total: float = Form(...)):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("UPDATE ventas SET id_producto=%s, cantidad=%s, precio_total=%s WHERE id_venta=%s", (id_producto, cantidad, precio_total, id_venta))
    conn.commit()
    cur.close()
    conn.close()
    return RedirectResponse(url="/ventas?msg=Venta+editada+con+éxito", status_code=303)

@app.post("/ventas/eliminar/{id_venta}")
def eliminar_venta(id_venta: int):
    conn = get_connection()
    cur = conn.cursor()
    cur.execute("DELETE FROM ventas WHERE id_venta=%s", (id_venta,))
    conn.commit()
    cur.close()
    conn.close()
    return RedirectResponse(url="/ventas?msg=Venta+eliminada+con+éxito", status_code=303)

@app.post("/ventas/nueva")
def registrar_venta(id_producto: int = Form(...), cantidad: int = Form(...), precio_total: float = Form(...)):
    insertar_venta(id_producto, cantidad, precio_total)
    return RedirectResponse(url="/ventas?msg=Venta+registrada+con+éxito", status_code=303)

@app.get("/ventas/exportar/excel")
def exportar_ventas_excel(filtro: str = ""):
    conn = get_connection()
    cur = conn.cursor()
    if filtro:
        cur.execute("""
            SELECT v.id_venta, p.nombre, v.cantidad, v.precio_total, v.fecha_venta
            FROM ventas v
            JOIN productos p ON v.id_producto = p.id_producto
            WHERE LOWER(p.nombre) LIKE %s
            ORDER BY v.id_venta DESC
        """, (f"%{filtro.lower()}%",))
    else:
        cur.execute("""
            SELECT v.id_venta, p.nombre, v.cantidad, v.precio_total, v.fecha_venta
            FROM ventas v
            JOIN productos p ON v.id_producto = p.id_producto
            ORDER BY v.id_venta DESC
        """)
    ventas = cur.fetchall()
    cur.close()
    conn.close()
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ID", "Producto", "Cantidad", "Precio Total", "Fecha"])
    for row in ventas:
        ws.append([row[0], row[1], row[2], row[3], row[4].strftime('%Y-%m-%d %H:%M') if row[4] else ''])
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet", headers={"Content-Disposition": "attachment; filename=ventas.xlsx"})

@app.get("/ventas/exportar/pdf")
def exportar_ventas_pdf(filtro: str = ""):
    conn = get_connection()
    cur = conn.cursor()
    if filtro:
        cur.execute("""
            SELECT v.id_venta, p.nombre, v.cantidad, v.precio_total, v.fecha_venta
            FROM ventas v
            JOIN productos p ON v.id_producto = p.id_producto
            WHERE LOWER(p.nombre) LIKE %s
            ORDER BY v.id_venta DESC
        """, (f"%{filtro.lower()}%",))
    else:
        cur.execute("""
            SELECT v.id_venta, p.nombre, v.cantidad, v.precio_total, v.fecha_venta
            FROM ventas v
            JOIN productos p ON v.id_producto = p.id_producto
            ORDER BY v.id_venta DESC
        """)
    ventas = cur.fetchall()
    cur.close()
    conn.close()
    pdf = FPDF()
    pdf.add_page()
    pdf.set_font("Arial", size=12)
    pdf.cell(0, 10, "Listado de Ventas", ln=True, align="C")
    pdf.ln(5)
    pdf.set_font("Arial", size=10)
    pdf.cell(15, 8, "ID", 1)
    pdf.cell(50, 8, "Producto", 1)
    pdf.cell(25, 8, "Cantidad", 1)
    pdf.cell(30, 8, "Precio Total", 1)
    pdf.cell(40, 8, "Fecha", 1)
    pdf.ln()
    for row in ventas:
        pdf.cell(15, 8, str(row[0]), 1)
        pdf.cell(50, 8, str(row[1]), 1)
        pdf.cell(25, 8, str(row[2]), 1)
        pdf.cell(30, 8, str(row[3]), 1)
        pdf.cell(40, 8, row[4].strftime('%Y-%m-%d %H:%M') if row[4] else '', 1)
        pdf.ln()
    stream = io.BytesIO(pdf.output(dest='S').encode('latin1'))
    stream.seek(0)
    return StreamingResponse(stream, media_type="application/pdf", headers={"Content-Disposition": "attachment; filename=ventas.pdf"})

@app.get("/stock-bajo")
def stock_bajo(request: Request):
    productos_bajos = consultar_productos_bajos()
    return templates.TemplateResponse("stock_bajo.html", {"request": request, "productos_bajos": productos_bajos})

@app.get("/categorias")
def categorias(request: Request, msg: str = Query(None)):
    categorias = listar_categorias()
    return templates.TemplateResponse("categorias.html", {"request": request, "categorias": categorias, "msg": msg})

@app.post("/categorias/nueva")
def categorias_nueva(nombre: str = Form(...), descripcion: str = Form("")):
    agregar_categoria(nombre, descripcion)
    return RedirectResponse(url="/categorias?msg=Categoría+agregada+con+éxito", status_code=303)

@app.post("/categorias/editar/{id_categoria}")
def categorias_editar(id_categoria: int, nombre: str = Form(...), descripcion: str = Form("")):
    editar_categoria(id_categoria, nombre, descripcion)
    return RedirectResponse(url="/categorias?msg=Categoría+editada+con+éxito", status_code=303)

@app.post("/categorias/eliminar/{id_categoria}")
def categorias_eliminar(id_categoria: int):
    eliminar_categoria(id_categoria)
    return RedirectResponse(url="/categorias?msg=Categoría+eliminada+con+éxito", status_code=303)

@app.get("/proveedores")
def proveedores(request: Request, msg: str = Query(None)):
    proveedores = listar_proveedores()
    return templates.TemplateResponse("proveedores.html", {"request": request, "proveedores": proveedores, "msg": msg})

@app.post("/proveedores/nuevo")
def proveedores_nuevo(nombre: str = Form(...), contacto: str = Form(""), telefono: str = Form(""), email: str = Form("")):
    agregar_proveedor(nombre, contacto, telefono, email)
    return RedirectResponse(url="/proveedores?msg=Proveedor+agregado+con+éxito", status_code=303)

@app.post("/proveedores/editar/{id_proveedor}")
def proveedores_editar(id_proveedor: int, nombre: str = Form(...), contacto: str = Form(""), telefono: str = Form(""), email: str = Form("")):
    editar_proveedor(id_proveedor, nombre, contacto, telefono, email)
    return RedirectResponse(url="/proveedores?msg=Proveedor+editado+con+éxito", status_code=303)

@app.post("/proveedores/eliminar/{id_proveedor}")
def proveedores_eliminar(id_proveedor: int):
    eliminar_proveedor(id_proveedor)
    return RedirectResponse(url="/proveedores?msg=Proveedor+eliminado+con+éxito", status_code=303)

@app.get("/usuarios")
def usuarios(request: Request, msg: str = Query(None)):
    return templates.TemplateResponse("usuarios.html", {"request": request, "msg": msg})

@app.post("/usuarios/registrar")
def usuarios_registrar(username: str = Form(...), password: str = Form(...)):
    registrar_usuario(username, password)
    return RedirectResponse(url="/usuarios?msg=Usuario+registrado+con+éxito", status_code=303)

@app.get("/logout")
def logout():
    return RedirectResponse(url="/login")

# Para desarrollo local
if __name__ == "__main__":
    uvicorn.run("app:app", host="127.0.0.1", port=8000, reload=True) 